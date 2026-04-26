"""
experiments/generate_plan.py
Generate experiments_plan.xlsx — full tracking sheet for all thesis runs.

Usage:
    python experiments/generate_plan.py
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parent / "experiments_plan.xlsx"

# ── Colour palette ────────────────────────────────────────────────────────────
COLOURS = {
    "B":      "D6EAF8",  # light blue    — internal baselines
    "B_COMBO":"85C1E9",  # medium blue   — standard combinations
    "B_PAPER":"A9CCE3",  # steel blue    — paper methods
    "M":      "D5F5E3",  # light green   — main CL runs
    "A_ETS":  "FEF9E7",  # light yellow  — ETS ablation
    "A_LPS":  "FDEBD0",  # light orange  — LPS ablation
    "A_MIX":  "F9EBEA",  # light red     — mixing ablation
    "A_LS":   "EBF5FB",  # pale cyan     — label smoothing
    "A_STR":  "F4ECF7",  # light purple  — strength
    "A_SCH":  "E8F8F5",  # mint          — LR scheduler
    "R":      "FDEDEC",  # blush         — seeds
    "D":      "D0ECE7",  # teal          — dataset generalization
    "F":      "E8DAEF",  # lavender      — final runs
    "HDR":    "2C3E50",  # dark navy     — header
}

STATUS_COLOURS = {
    "Done":    "27AE60",
    "Running": "F39C12",
    "Failed":  "E74C3C",
    "Pending": "BDC3C7",
}

# ── Experiment table ──────────────────────────────────────────────────────────
# Columns: ID | Group | Name | Description | Aug | Schedule | Extra Config |
#          Mix Mode | Seed | Epochs | Status | Val Acc | Test Top-1 |
#          Test Top-5 | Best Epoch | Runtime | Notes

EXPERIMENTS = [

    # ── B — Internal baselines ────────────────────────────────────────────────
    ("B1", "B", "No Augmentation",
     "Trains with zero data augmentation. Sets the absolute performance floor "
     "to show how much all augmentation methods gain over a vanilla model.",
     "none", "—", "—", "—", 42, 100, "Lower bound"),

    ("B2", "B", "Static Augmentation",
     "Applies the full Tier-3 augmentation pool at a fixed strength from epoch 1 "
     "with no schedule. Isolates the value of the curriculum itself — if CL beats "
     "this, the progressive ordering is contributing, not just the op set.",
     "static", "—", "—", "—", 42, 100, "Key non-CL comparison"),

    ("B3", "B", "Random Augmentation",
     "Same op pool as the curriculum but each op is applied with an independently "
     "sampled random strength every call. Tests whether controlled strength "
     "ordering matters versus pure randomness.",
     "random", "—", "—", "—", 42, 100, ""),

    ("B4", "B", "RandAugment N=2 M=9",
     "Official RandAugment (Cubuk et al. NeurIPS 2020) with CIFAR-100 defaults. "
     "The most widely used automatic augmentation baseline — essential for "
     "situating results in the literature.",
     "randaugment", "—", "N=2 M=9", "—", 42, 100, "Cubuk et al. NeurIPS 2020"),

    ("B5", "B", "Static + Mixing",
     "Static augmentation pipeline with CutMix and MixUp applied from epoch 1. "
     "Isolates the effect of adding batch mixing to a non-curriculum baseline, "
     "separating mixing gains from curriculum gains.",
     "static_mixing", "—", "—", "both", 42, 100, ""),

    # ── B_COMBO — Standard manual pipeline baselines ──────────────────────────
    ("B_COMBO1", "B_COMBO", "Flip + Crop",
     "RandomHorizontalFlip + RandomCrop(32, pad=4). The de-facto standard "
     "baseline used in the original ResNet paper and almost every CIFAR-100 "
     "result table. Essential reference point.",
     "flip_crop", "—", "pad=4", "—", 42, 100, "He et al. ResNet 2016"),

    ("B_COMBO2", "B_COMBO", "Flip + Crop + ColorJitter",
     "Adds colour distortion (brightness, contrast, saturation, hue) on top of "
     "the standard flip+crop. Represents a typical strong manual pipeline used "
     "in many supervised learning papers.",
     "flip_crop_color", "—", "b=0.4 c=0.4 s=0.4", "—", 42, 100, ""),

    ("B_COMBO3", "B_COMBO", "Flip + Crop + CutOut",
     "Standard flip+crop base augmented with random patch erasing (CutOut). "
     "Tests whether structured information removal alone, without a curriculum, "
     "yields competitive accuracy.",
     "flip_crop_cutout", "—", "size=16px", "—", 42, 100, "DeVries & Taylor 2017"),

    ("B_COMBO4", "B_COMBO", "Flip + Crop + Color + CutOut",
     "Combines spatial, colour, and erasing transforms into one strong manual "
     "pipeline. Commonly appears as the non-automated baseline in SOTA comparison "
     "tables for CIFAR-100.",
     "flip_crop_color_cut", "—", "—", "—", 42, 100, "Strong manual baseline"),

    ("B_COMBO5", "B_COMBO", "Flip + Crop + MixUp",
     "Standard spatial base plus MixUp (Zhang et al. ICLR 2018). Establishes "
     "what mixing alone contributes on top of the minimal pipeline, independent "
     "of any curriculum structure.",
     "flip_crop_mixup", "—", "alpha=1.0", "mixup", 42, 100, "Zhang et al. ICLR 2018"),

    ("B_COMBO6", "B_COMBO", "Flip + Crop + CutMix",
     "Standard spatial base plus CutMix (Yun et al. ICCV 2019). Compared with "
     "B_COMBO5 this shows whether patch-based mixing outperforms linear "
     "interpolation on a minimal pipeline.",
     "flip_crop_cutmix", "—", "alpha=1.0", "cutmix", 42, 100, "Yun et al. ICCV 2019"),

    ("B_COMBO7", "B_COMBO", "Flip + Crop + CutMix + MixUp",
     "Combines both mixing strategies on top of the standard base. Shows whether "
     "mixing both modes jointly improves over either alone before adding the "
     "curriculum component.",
     "flip_crop_both_mix", "—", "alpha=1.0", "both", 42, 100, ""),

    ("B_COMBO8", "B_COMBO", "Tier-1 Ops Fixed (all epochs)",
     "Applies only the Tier-1 augmentation set (flip, crop, translate x/y) for "
     "all 100 epochs without advancing. Upper-bounds what Tier-1 alone can "
     "achieve and motivates the need for harder tiers.",
     "tier1_static", "—", "flip+crop+translate", "—", 42, 100, ""),

    ("B_COMBO9", "B_COMBO", "Tier-2 Ops Fixed (all epochs)",
     "Applies the Tier-2 op set (+ colour jitter, rotation, shear, equalize, "
     "sharpness, auto-contrast) for all 100 epochs. Compared with Tier-1 static "
     "and the full curriculum it shows the contribution of each tier boundary.",
     "tier2_static", "—", "+color+rot+shear", "—", 42, 100, ""),

    # ── B_PAPER — Published method baselines ──────────────────────────────────
    ("B_PAPER1", "B_PAPER", "AutoAugment CIFAR-100",
     "Learns a dataset-specific augmentation policy via reinforcement learning "
     "(Cubuk et al. CVPR 2019). The first automated augmentation method and still "
     "a standard reference in augmentation papers.",
     "autoaugment", "—", "torchvision built-in", "—", 42, 100, "Cubuk et al. CVPR 2019"),

    ("B_PAPER2", "B_PAPER", "TrivialAugment",
     "Samples one random operation at a uniformly random magnitude each call "
     "(Müller & Hutter ICCV 2021). Despite its simplicity it matches or beats "
     "AutoAugment — important to include as a strong cheap baseline.",
     "trivialaugment", "—", "torchvision built-in", "—", 42, 100, "Müller & Hutter ICCV 2021"),

    ("B_PAPER3", "B_PAPER", "AugMix",
     "Mixes multiple augmented views of each image to improve robustness and "
     "uncertainty calibration (Hendrycks et al. ICLR 2020). Relevant because "
     "it also applies augmentations at varying intensities, making it a natural "
     "curriculum-adjacent comparison.",
     "augmix", "—", "torchvision built-in", "—", 42, 100, "Hendrycks et al. ICLR 2020"),

    ("B_PAPER4", "B_PAPER", "CutOut / RandomErasing",
     "Randomly masks out a rectangular patch of each image (DeVries & Taylor "
     "2017). A simple single-op erasing baseline used across many CIFAR-100 "
     "tables and part of our Tier-3 op set.",
     "cutout_erasing", "—", "torchvision built-in", "—", 42, 100, "DeVries & Taylor 2017"),

    ("B_PAPER5", "B_PAPER", "MixUp (standalone)",
     "Applies linear interpolation between random image pairs and their labels "
     "from the very first epoch (Zhang et al. ICLR 2018). Baseline for "
     "comparing against curriculum-gated mixing in Tier 3.",
     "mixup_standalone", "—", "alpha=1.0 p=0.5", "mixup", 42, 100, "Zhang et al. ICLR 2018"),

    ("B_PAPER6", "B_PAPER", "CutMix (standalone)",
     "Cuts and pastes image patches between random pairs with mixed labels from "
     "epoch 1 (Yun et al. ICCV 2019). Baseline for comparing against "
     "curriculum-gated CutMix introduced only in Tier 3.",
     "cutmix_standalone", "—", "alpha=1.0 p=0.5", "cutmix", 42, 100, "Yun et al. ICCV 2019"),

    ("B_PAPER7", "B_PAPER", "MADAug",
     "Data-driven adaptive augmentation that adjusts per-sample augmentation "
     "strength based on model confidence (Park et al.). Already present in "
     "augmentations/sota/MADAug/ — the most direct competing curriculum method.",
     "madaug", "—", "sota/MADAug/ in repo", "—", 42, 100, "Park et al."),

    ("B_PAPER8", "B_PAPER", "Fast AutoAugment",
     "Replaces the RL search of AutoAugment with density matching for a 100× "
     "speedup (Lim et al. NeurIPS 2019). Tests whether a faster searched policy "
     "is competitive with the hand-designed curriculum.",
     "fast_autoaugment", "—", "pip fastautoaugment", "—", 42, 100, "Lim et al. NeurIPS 2019"),

    ("B_PAPER9", "B_PAPER", "GridMask",
     "Removes structured grid-shaped regions from images to encourage the model "
     "to use non-local features (Chen et al. 2020). A structured erasing "
     "alternative to the random CutOut used in Tier 3.",
     "gridmask", "—", "~30-line wrapper", "—", 42, 100, "Chen et al. 2020"),

    # ── M — Main CL runs ──────────────────────────────────────────────────────
    ("M1", "M", "ETS — No Mixing",
     "Three-tier curriculum with epoch-threshold scheduling and no batch mixing. "
     "Isolates the contribution of the progressive augmentation schedule alone, "
     "before mixing is added.",
     "tiered_curriculum", "ets", "—", "none", 42, 100, "ETS lower bound"),

    ("M2", "M", "ETS — CutMix only",
     "ETS curriculum with CutMix applied in Tier 3. Tests whether patch-based "
     "mixing in the hardest tier provides a consistent gain over the schedule "
     "alone.",
     "tiered_curriculum", "ets", "—", "cutmix", 42, 100, ""),

    ("M3", "M", "ETS — MixUp only",
     "ETS curriculum with MixUp applied in Tier 3. Compared with M2 this "
     "reveals whether linear interpolation or patch mixing is the better "
     "complement to the ETS schedule.",
     "tiered_curriculum", "ets", "—", "mixup", 42, 100, ""),

    ("M4", "M", "ETS — CutMix + MixUp",
     "Primary ETS entry. Epoch-threshold scheduling with both mixing strategies "
     "active in Tier 3. The main fixed-schedule variant reported in the thesis "
     "comparison table.",
     "tiered_curriculum", "ets", "—", "both", 42, 100, "Primary ETS ★"),

    ("M5", "M", "LPS — No Mixing",
     "Three-tier curriculum with loss-plateau scheduling and no batch mixing. "
     "Establishes the LPS baseline before mixing is added, enabling a clean "
     "schedule-vs-mixing decomposition.",
     "tiered_curriculum", "lps", "—", "none", 42, 100, "LPS lower bound"),

    ("M6", "M", "LPS — CutMix only",
     "LPS curriculum with CutMix in Tier 3. Tests whether adaptive tier "
     "advancement combined with patch mixing outperforms fixed-schedule CutMix "
     "(M2).",
     "tiered_curriculum", "lps", "—", "cutmix", 42, 100, ""),

    ("M7", "M", "LPS — MixUp only",
     "LPS curriculum with MixUp in Tier 3. Compared with M6 this reveals "
     "whether adaptive scheduling changes which mixing mode is preferable.",
     "tiered_curriculum", "lps", "—", "mixup", 42, 100, ""),

    ("M8", "M", "LPS — CutMix + MixUp",
     "Primary LPS entry. Loss-plateau scheduling with both mixing strategies "
     "active in Tier 3. The adaptive-schedule variant reported alongside M4 as "
     "the core thesis comparison. Currently running.",
     "tiered_curriculum", "lps", "—", "both", 42, 100, "Primary LPS ★ — running"),

    # ── A_ETS — Tier boundary ablation ───────────────────────────────────────
    ("A_ETS1", "A_ETS", "ETS t1=0.15 t2=0.40",
     "Tests earlier tier transitions (Tier 1 ends at 15%, Tier 3 starts at 40%). "
     "Answers whether advancing the curriculum sooner hurts because the model has "
     "not yet learned basic features.",
     "tiered_curriculum", "ets", "t1=0.15 t2=0.40", "both", 42, 100, "Earlier schedule"),

    ("A_ETS2", "A_ETS", "ETS t1=0.20 t2=0.45",
     "Optimized tier boundaries found by grid search. Tier 1 runs for 20% of "
     "epochs, Tier 2 for 25%, Tier 3 for 55%. Used as the default in all main "
     "ETS runs.",
     "tiered_curriculum", "ets", "t1=0.20 t2=0.45", "both", 42, 100, "Optimized default ★"),

    ("A_ETS3", "A_ETS", "ETS t1=0.25 t2=0.50",
     "Slightly later transitions than the optimized default. Tests whether "
     "spending more time in Tier 1 before introducing harder ops improves "
     "generalization.",
     "tiered_curriculum", "ets", "t1=0.25 t2=0.50", "both", 42, 100, ""),

    ("A_ETS4", "A_ETS", "ETS t1=0.33 t2=0.66",
     "Equal thirds — the naive default split where each tier gets the same "
     "number of epochs. Included to show why optimized boundaries matter over "
     "the obvious uniform division.",
     "tiered_curriculum", "ets", "t1=0.33 t2=0.66", "both", 42, 100, "Equal-thirds naive split"),

    ("A_ETS5", "A_ETS", "ETS t1=0.20 t2=0.55",
     "Extends Tier 2 relative to the default (0.45 → 0.55) while keeping Tier 1 "
     "the same. Tests whether the model benefits from a longer intermediate "
     "augmentation phase before the hardest ops.",
     "tiered_curriculum", "ets", "t1=0.20 t2=0.55", "both", 42, 100, "Longer Tier 2"),

    ("A_ETS6", "A_ETS", "ETS t1=0.30 t2=0.60",
     "Later transitions throughout — more conservative schedule. Tests the "
     "opposite hypothesis from A_ETS1: does starting harder augmentations later "
     "allow the model to build a stronger foundation first?",
     "tiered_curriculum", "ets", "t1=0.30 t2=0.60", "both", 42, 100, "Conservative schedule"),

    # ── A_LPS — LPS hyperparameter ablation ──────────────────────────────────
    ("A_LPS1", "A_LPS", "LPS tau=0.005",
     "Very sensitive plateau threshold — advances tier only when improvement is "
     "nearly zero. May stay in a tier too long if loss oscillates slightly above "
     "threshold.",
     "tiered_curriculum", "lps", "tau=0.005 win=5 minep=10", "both", 42, 100, ""),

    ("A_LPS2", "A_LPS", "LPS tau=0.010",
     "Stricter than default tau. Requires 1% relative improvement in the loss "
     "window to stay in the current tier. Tests whether tighter plateau detection "
     "advances the schedule more efficiently.",
     "tiered_curriculum", "lps", "tau=0.010 win=5 minep=10", "both", 42, 100, ""),

    ("A_LPS3", "A_LPS", "LPS tau=0.020",
     "Default plateau threshold (2% relative improvement). Validated on the "
     "development run. All other LPS ablations hold this fixed unless otherwise "
     "stated.",
     "tiered_curriculum", "lps", "tau=0.020 win=5 minep=10", "both", 42, 100, "Default ★"),

    ("A_LPS4", "A_LPS", "LPS tau=0.050",
     "Looser threshold — advances tier if improvement drops below 5%. Will "
     "trigger tier advances earlier, potentially before the model has fully "
     "exploited the current augmentation level.",
     "tiered_curriculum", "lps", "tau=0.050 win=5 minep=10", "both", 42, 100, ""),

    ("A_LPS5", "A_LPS", "LPS tau=0.100",
     "Aggressive threshold — triggers tier advance when any non-trivial "
     "improvement stalls. Expected to advance too quickly, acting almost like "
     "an early fixed schedule.",
     "tiered_curriculum", "lps", "tau=0.100 win=5 minep=10", "both", 42, 100, "Aggressive"),

    ("A_LPS6", "A_LPS", "LPS window=3",
     "Shorter look-back window for plateau detection. Averaging over only 3 "
     "epochs makes the signal noisier and may trigger false advances due to "
     "per-epoch loss fluctuations.",
     "tiered_curriculum", "lps", "tau=0.020 win=3 minep=10", "both", 42, 100, "Short window"),

    ("A_LPS7", "A_LPS", "LPS window=7",
     "Slightly longer look-back than default. Smooths out more noise before "
     "deciding to advance, potentially delaying tier transitions by 2–4 epochs "
     "compared to win=5.",
     "tiered_curriculum", "lps", "tau=0.020 win=7 minep=10", "both", 42, 100, ""),

    ("A_LPS8", "A_LPS", "LPS window=10",
     "Long look-back window. The plateau signal becomes very stable but slow to "
     "react. Tests whether stable but delayed detection leads to better or worse "
     "final accuracy than the responsive default.",
     "tiered_curriculum", "lps", "tau=0.020 win=10 minep=10", "both", 42, 100, "Long window"),

    ("A_LPS9", "A_LPS", "LPS min_epochs=5",
     "Allows tier advancement after only 5 epochs minimum per tier. Risks "
     "advancing before the model has stabilized on the current augmentation "
     "difficulty level.",
     "tiered_curriculum", "lps", "tau=0.020 win=5 minep=5", "both", 42, 100, "Short min stay"),

    ("A_LPS10", "A_LPS", "LPS min_epochs=15",
     "Forces the model to spend at least 15 epochs in each tier before any "
     "plateau check. Prevents premature advancement while still allowing "
     "adaptive transitions.",
     "tiered_curriculum", "lps", "tau=0.020 win=5 minep=15", "both", 42, 100, ""),

    ("A_LPS11", "A_LPS", "LPS min_epochs=20",
     "Long mandatory stay per tier (20 epochs). At 100 total epochs this "
     "guarantees at least 60 epochs are locked into Tier 1 and 2, leaving at "
     "most 60 for Tier 3. Tests whether forced patience improves stability.",
     "tiered_curriculum", "lps", "tau=0.020 win=5 minep=20", "both", 42, 100, "Long min stay"),

    # ── A_MIX — Mixing ablation ───────────────────────────────────────────────
    ("A_MIX1", "A_MIX", "ETS — No Mixing",
     "ETS curriculum with mixing completely disabled. The control condition for "
     "the mixing ablation group — establishes what the augmentation schedule "
     "contributes without any batch mixing.",
     "tiered_curriculum", "ets", "—", "none", 42, 100, "Control"),

    ("A_MIX2", "A_MIX", "ETS — CutMix only",
     "Patch-based mixing only in Tier 3. Compared with A_MIX1 this isolates the "
     "contribution of CutMix to the final accuracy.",
     "tiered_curriculum", "ets", "—", "cutmix", 42, 100, ""),

    ("A_MIX3", "A_MIX", "ETS — MixUp only",
     "Linear interpolation mixing only in Tier 3. Compared with A_MIX2 reveals "
     "whether the type of mixing (patch vs blend) matters for this curriculum.",
     "tiered_curriculum", "ets", "—", "mixup", 42, 100, ""),

    ("A_MIX4", "A_MIX", "ETS — Both modes",
     "Default mixing configuration: both CutMix and MixUp randomly applied in "
     "Tier 3. Reference point for all other mixing ablations.",
     "tiered_curriculum", "ets", "—", "both", 42, 100, "Default ★"),

    ("A_MIX5", "A_MIX", "ETS — Both + ramp",
     "Gradually ramps mixing probability and alpha from 0 to full strength over "
     "the first 5 epochs of Tier 3. Tests whether a smooth mixing introduction "
     "reduces the accuracy drop at the Tier 3 boundary.",
     "tiered_curriculum", "ets", "mix_ramp=True", "both", 42, 100, ""),

    ("A_MIX6", "A_MIX", "ETS — mix_prob=0.3",
     "Reduces mixing frequency to 30% of batches in Tier 3. Tests whether less "
     "frequent mixing is sufficient or whether higher coverage is important for "
     "regularization.",
     "tiered_curriculum", "ets", "mix_prob=0.3", "both", 42, 100, "Low probability"),

    ("A_MIX7", "A_MIX", "ETS — mix_prob=0.7",
     "Increases mixing frequency to 70% of batches. Tests whether applying "
     "mixing to more batches per epoch further improves generalization or causes "
     "training instability.",
     "tiered_curriculum", "ets", "mix_prob=0.7", "both", 42, 100, "High probability"),

    ("A_MIX8", "A_MIX", "ETS — mix_alpha=0.2",
     "Low Beta distribution parameter — produces mix ratios strongly skewed "
     "toward 0 or 1, making most mixed images look nearly like one source image. "
     "Tests weak mixing.",
     "tiered_curriculum", "ets", "mix_alpha=0.2", "both", 42, 100, "Weak mix ratio"),

    ("A_MIX9", "A_MIX", "ETS — mix_alpha=0.5",
     "Moderate Beta parameter — produces a mix ratio distribution that peaks "
     "near 0 and 1 but with more mid-range samples than alpha=0.2. A common "
     "alternative to the default alpha=1.0.",
     "tiered_curriculum", "ets", "mix_alpha=0.5", "both", 42, 100, ""),

    ("A_MIX10", "A_MIX", "ETS — mix_alpha=2.0",
     "High Beta parameter — produces mix ratios concentrated near 0.5, creating "
     "images that are roughly equal blends of two sources. Tests aggressive "
     "mixing compared to the uniform default.",
     "tiered_curriculum", "ets", "mix_alpha=2.0", "both", 42, 100, "Strong mix ratio"),

    # ── A_LS — Label smoothing ablation ──────────────────────────────────────
    ("A_LS1", "A_LS", "ETS — no label smoothing",
     "Hard one-hot targets with no smoothing. The control for the label "
     "smoothing ablation. Used in all main runs as the default.",
     "tiered_curriculum", "ets", "smooth=0.00", "both", 42, 100, "Default ★"),

    ("A_LS2", "A_LS", "ETS — label smooth=0.05",
     "Mild label smoothing. Each incorrect class receives 0.05/99 of the "
     "probability mass. Tests whether any smoothing helps on CIFAR-100 with the "
     "curriculum without over-regularizing.",
     "tiered_curriculum", "ets", "smooth=0.05", "both", 42, 100, ""),

    ("A_LS3", "A_LS", "ETS — label smooth=0.10",
     "Standard label smoothing coefficient (ε=0.1) commonly used in "
     "Transformer and large-model training. Most likely to interact positively "
     "with the hard augmentations in Tier 3.",
     "tiered_curriculum", "ets", "smooth=0.10", "both", 42, 100, ""),

    ("A_LS4", "A_LS", "ETS — label smooth=0.15",
     "Stronger smoothing that distributes more probability mass to wrong classes. "
     "Tests whether heavy smoothing combined with aggressive Tier-3 augmentation "
     "hurts or helps calibration.",
     "tiered_curriculum", "ets", "smooth=0.15", "both", 42, 100, "Strong smoothing"),

    ("A_LS5", "A_LS", "LPS — label smooth=0.10",
     "Applies the best smoothing value from the ETS ablation to the LPS "
     "schedule. Checks whether the benefit of smoothing transfers across "
     "scheduling mechanisms.",
     "tiered_curriculum", "lps", "smooth=0.10", "both", 42, 100, ""),

    # ── A_STR — Augmentation strength ablation ───────────────────────────────
    ("A_STR1", "A_STR", "ETS — strength=0.5",
     "Lowers the augmentation strength ceiling to 50%. Tiers still ramp from "
     "40% to 100% of this ceiling, so the absolute max intensity is half the "
     "default. Tests if lighter augmentation helps on CIFAR-100.",
     "tiered_curriculum", "ets", "strength=0.5", "both", 42, 100, "Weak augmentation"),

    ("A_STR2", "A_STR", "ETS — strength=0.7",
     "Default augmentation strength ceiling used in all main runs. The "
     "reference point for the strength ablation.",
     "tiered_curriculum", "ets", "strength=0.7", "both", 42, 100, "Default ★"),

    ("A_STR3", "A_STR", "ETS — strength=0.9",
     "Increases the strength ceiling to 90%. The model sees near-maximum "
     "distortion in Tier 3. Tests whether stronger augmentation at the cost of "
     "harder training improves regularization.",
     "tiered_curriculum", "ets", "strength=0.9", "both", 42, 100, "Strong augmentation"),

    ("A_STR4", "A_STR", "LPS — strength=0.5",
     "Weak augmentation with adaptive scheduling. Tests whether LPS compensates "
     "for lighter augmentation by spending longer in each tier, or whether the "
     "weaker signal limits the final accuracy regardless.",
     "tiered_curriculum", "lps", "strength=0.5", "both", 42, 100, ""),

    ("A_STR5", "A_STR", "LPS — strength=0.9",
     "Strong augmentation with adaptive scheduling. The most aggressive "
     "setting — tests whether LPS can handle very hard augmentations by "
     "staying in easier tiers until the model is ready.",
     "tiered_curriculum", "lps", "strength=0.9", "both", 42, 100, ""),

    # ── A_SCH — LR scheduler ablation ────────────────────────────────────────
    ("A_SCH1", "A_SCH", "ETS — MultiStepLR",
     "Default LR schedule: drops by 10× at 33%, 66%, and 83% of total epochs. "
     "The reference point for the scheduler ablation — milestone drops align "
     "with tier transitions in ETS.",
     "tiered_curriculum", "ets", "scheduler=multistep", "both", 42, 100, "Default ★"),

    ("A_SCH2", "A_SCH", "ETS — CosineAnnealingLR",
     "Replaces step-wise LR drops with a smooth cosine decay. Tests whether "
     "the abrupt LR drops that reset LPS plateau history are necessary, or "
     "whether a smooth schedule is better suited to the curriculum.",
     "tiered_curriculum", "ets", "scheduler=cosine", "both", 42, 100, ""),

    ("A_SCH3", "A_SCH", "LPS — MultiStepLR",
     "LPS with the default MultiStep LR schedule. LR drops trigger a plateau "
     "history reset in the scheduler, preventing false advances after the loss "
     "temporarily spikes.",
     "tiered_curriculum", "lps", "scheduler=multistep", "both", 42, 100, ""),

    ("A_SCH4", "A_SCH", "LPS — CosineAnnealingLR",
     "LPS with cosine LR decay. Without step drops there are no grace-period "
     "resets, so the plateau detector sees a monotonically decaying loss. Tests "
     "whether this cleaner signal improves tier-advance timing.",
     "tiered_curriculum", "lps", "scheduler=cosine", "both", 42, 100, ""),

    ("A_SCH5", "A_SCH", "Static — CosineAnnealingLR",
     "Static augmentation with cosine LR decay. Isolates the LR schedule effect "
     "on the non-CL baseline to ensure any difference seen in A_SCH2/4 is "
     "due to the curriculum interaction, not the scheduler alone.",
     "static", "—", "scheduler=cosine", "—", 42, 100, ""),

    # ── R — Multi-seed robustness ─────────────────────────────────────────────
    ("R1", "R", "ETS — seed 42",
     "Primary ETS run (seed=42). Reference for the seed robustness group.",
     "tiered_curriculum", "ets", "—", "both", 42, 100, "Reference seed"),

    ("R2", "R", "ETS — seed 123",
     "Second ETS run with different initialisation and data order. Together with "
     "R1, R3, R4 estimates ETS variance and confirms results are not seed-lucky.",
     "tiered_curriculum", "ets", "—", "both", 123, 100, ""),

    ("R3", "R", "ETS — seed 7",
     "Third ETS seed. Three seeds are the minimum for a mean ± std report "
     "in a thesis comparison table.",
     "tiered_curriculum", "ets", "—", "both", 7, 100, ""),

    ("R4", "R", "ETS — seed 2024",
     "Fourth ETS seed. Provides tighter confidence interval for the ETS mean "
     "accuracy reported in the thesis.",
     "tiered_curriculum", "ets", "—", "both", 2024, 100, ""),

    ("R5", "R", "LPS — seed 42",
     "Primary LPS run (seed=42). Reference for the LPS seed robustness group.",
     "tiered_curriculum", "lps", "—", "both", 42, 100, "Reference seed"),

    ("R6", "R", "LPS — seed 123",
     "Second LPS seed. LPS tier transitions depend on the loss curve, which "
     "varies with initialization — multiple seeds show how much the transition "
     "timing varies.",
     "tiered_curriculum", "lps", "—", "both", 123, 100, ""),

    ("R7", "R", "LPS — seed 7",
     "Third LPS seed.",
     "tiered_curriculum", "lps", "—", "both", 7, 100, ""),

    ("R8", "R", "LPS — seed 2024",
     "Fourth LPS seed. Enables reporting LPS mean ± std alongside ETS for a "
     "fair statistical comparison in the thesis.",
     "tiered_curriculum", "lps", "—", "both", 2024, 100, ""),

    ("R9", "R", "Static — seed 123",
     "Second seed for the static baseline. Confirms that the baseline variance "
     "is lower than the CL variance, showing the schedule introduces "
     "additional but beneficial stochasticity.",
     "static", "—", "—", "—", 123, 100, ""),

    ("R10", "R", "RandAugment — seed 123",
     "Second seed for RandAugment. Needed to report RandAugment with error bars "
     "in the comparison table alongside the curriculum results.",
     "randaugment", "—", "N=2 M=9", "—", 123, 100, ""),

    # ── D — Dataset generalization ────────────────────────────────────────────
    ("D1", "D", "Tiny ImageNet — No Aug",
     "No-augmentation lower bound on Tiny ImageNet (200 classes, 64×64). "
     "Establishes the performance floor on a harder dataset to test whether "
     "the curriculum generalizes beyond CIFAR-100.",
     "none", "—", "—", "—", 42, 100, "Tiny ImageNet lower bound"),

    ("D2", "D", "Tiny ImageNet — Static",
     "Static augmentation on Tiny ImageNet. The non-CL baseline for the "
     "generalization group — shows whether the augmentation ops transfer to "
     "higher-resolution images.",
     "static", "—", "—", "—", 42, 100, ""),

    ("D3", "D", "Tiny ImageNet — RandAugment",
     "RandAugment on Tiny ImageNet. A widely reported paper baseline for this "
     "dataset — needed to check that our CL approach is competitive on a second "
     "benchmark.",
     "randaugment", "—", "N=2 M=9", "—", 42, 100, ""),

    ("D4", "D", "Tiny ImageNet — ETS + Mixing",
     "Best ETS configuration transferred to Tiny ImageNet. Tests whether the "
     "tier boundaries and mixing strategy optimized for CIFAR-100 carry over "
     "without retuning.",
     "tiered_curriculum", "ets", "—", "both", 42, 100, ""),

    ("D5", "D", "Tiny ImageNet — LPS + Mixing",
     "Best LPS configuration transferred to Tiny ImageNet. Because LPS adapts "
     "to the loss curve rather than fixed fractions, it may generalize better "
     "than ETS to a new dataset.",
     "tiered_curriculum", "lps", "—", "both", 42, 100, ""),

    # ── F — Final full-train runs ─────────────────────────────────────────────
    ("F1", "F", "Final — No Augmentation",
     "Full-train run (val_split=0.0, all 50k samples) with no augmentation. "
     "The final lower-bound number reported in the thesis test-set table.",
     "none", "—", "—", "—", 42, 100, "val_split=0.0"),

    ("F2", "F", "Final — Static",
     "Full-train static augmentation. The definitive non-CL baseline number for "
     "the thesis — trained on the full training set so the comparison is fair.",
     "static", "—", "—", "—", 42, 100, "val_split=0.0"),

    ("F3", "F", "Final — RandAugment",
     "Full-train RandAugment. The published-method comparison number for the "
     "thesis table — trained on all 50k samples.",
     "randaugment", "—", "N=2 M=9", "—", 42, 100, "val_split=0.0"),

    ("F4", "F", "Final — ETS + Mixing",
     "Full-train primary ETS run. This is the definitive ETS thesis number. "
     "Reported alongside F5 as the main result of the thesis.",
     "tiered_curriculum", "ets", "—", "both", 42, 100, "val_split=0.0  ← thesis number ★"),

    ("F5", "F", "Final — LPS + Mixing",
     "Full-train primary LPS run. The definitive LPS thesis number. The core "
     "claim of the thesis rests on whether this outperforms F4 and the "
     "baselines.",
     "tiered_curriculum", "lps", "—", "both", 42, 100, "val_split=0.0  ← thesis number ★"),

    ("F6", "F", "Final — Static + Mixing",
     "Full-train static augmentation with both mixing strategies from epoch 1. "
     "Separates the contribution of mixing from the curriculum in the final "
     "results table.",
     "static_mixing", "—", "—", "both", 42, 100, "val_split=0.0"),

    ("F7", "F", "Final — AutoAugment",
     "Full-train AutoAugment. The strongest paper-baseline final number — "
     "necessary to show the curriculum competes with RL-searched policies.",
     "autoaugment", "—", "—", "—", 42, 100, "val_split=0.0"),

    ("F8", "F", "Final — TrivialAugment",
     "Full-train TrivialAugment. The simplest strong baseline final number — "
     "if the curriculum cannot beat TrivialAugment on the full training set, "
     "the added complexity is hard to justify.",
     "trivialaugment", "—", "—", "—", 42, 100, "val_split=0.0"),
]

# ── Sheet layout ──────────────────────────────────────────────────────────────
HEADERS = [
    "ID", "Group", "Experiment Name", "Description",
    "Augmentation", "Schedule", "Extra Config", "Mix Mode", "Seed", "Epochs",
    "Status", "Val Acc (%)", "Test Top-1 (%)", "Test Top-5 (%)",
    "Best Epoch", "Runtime (min)", "Notes",
]

COL_WIDTHS = [10, 9, 28, 55, 18, 10, 22, 10, 7, 8, 10, 12, 13, 13, 11, 14, 30]


def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour)


def build_sheet(wb: openpyxl.Workbook):
    ws = wb.active
    ws.title = "Experiment Tracker"

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = _fill(COLOURS["HDR"])
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = _border()
    ws.row_dimensions[1].height = 32

    for row_idx, exp in enumerate(EXPERIMENTS, 2):
        (exp_id, group, name, description,
         aug, schedule, extra, mix, seed, epochs, notes) = exp

        row_data = [
            exp_id, group, name, description,
            aug, schedule, extra, mix, seed, epochs,
            "Pending", "", "", "", "", "", notes,
        ]

        fill  = _fill(COLOURS.get(group, "FFFFFF"))
        align = Alignment(vertical="top", wrap_text=True)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.alignment = align
            cell.border    = _border()
            if col_idx == 1:
                cell.font = Font(bold=True)
            if col_idx == 11:
                cell.alignment = Alignment(horizontal="center", vertical="top")

        ws.row_dimensions[row_idx].height = 60

    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


def build_legend(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Legend")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["D"].width = 12

    ws.cell(1, 1, "Group").font     = Font(bold=True)
    ws.cell(1, 2, "Description").font = Font(bold=True)
    ws.cell(1, 4, "Status").font    = Font(bold=True)

    group_desc = [
        ("B",       "Internal baselines — no aug / static / random / RandAugment / static+mixing"),
        ("B_COMBO", "Manual pipeline baselines — standard combinations used in CIFAR papers"),
        ("B_PAPER", "Published method baselines — AutoAugment, TrivialAugment, AugMix, etc."),
        ("M",       "Main CL runs — ETS vs LPS × mix_mode  (core thesis comparison)"),
        ("A_ETS",   "Ablation — ETS tier-boundary fractions (tier_t1, tier_t2)"),
        ("A_LPS",   "Ablation — LPS hyperparameters (tau / window / min_epochs_per_tier)"),
        ("A_MIX",   "Ablation — batch mixing mode, probability, alpha, ramp"),
        ("A_LS",    "Ablation — label smoothing coefficient"),
        ("A_STR",   "Ablation — augmentation op strength ceiling"),
        ("A_SCH",   "Ablation — LR scheduler (MultiStep vs CosineAnnealing)"),
        ("R",       "Robustness — multiple random seeds for mean ± std reporting"),
        ("D",       "Dataset generalization — Tiny ImageNet (200 classes, 64×64)"),
        ("F",       "Final full-train runs (val_split=0.0) — definitive thesis numbers"),
    ]
    for i, (grp, desc) in enumerate(group_desc, 2):
        ws.cell(i, 1, grp).fill = _fill(COLOURS[grp])
        ws.cell(i, 1, grp).font = Font(bold=True)
        ws.cell(i, 2, desc)
        ws.row_dimensions[i].height = 18

    for i, (status, colour) in enumerate(STATUS_COLOURS.items(), 2):
        cell = ws.cell(i, 4, status)
        cell.fill      = _fill(colour)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")


def main():
    wb = openpyxl.Workbook()
    build_sheet(wb)
    build_legend(wb)
    wb.save(OUTPUT)
    groups = {}
    for exp in EXPERIMENTS:
        groups[exp[1]] = groups.get(exp[1], 0) + 1
    print(f"Saved : {OUTPUT}")
    print(f"Total : {len(EXPERIMENTS)} experiments")
    for grp, count in groups.items():
        print(f"  {grp:<9} {count:>3} runs")


if __name__ == "__main__":
    main()
